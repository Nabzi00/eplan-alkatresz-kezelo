import streamlit as st
import pandas as pd
import io
import openpyxl
import json
import os
import datetime
import streamlit.components.v1 as components

# Oldal beállításai
st.set_page_config(page_title="Ekszi segéd", layout="wide", initial_sidebar_state="expanded")

# --- CSS: CÍM FELHÚZÁSA ÉS CACODEMON EASTER EGG ---
st.markdown("""
    <style>
    .block-container { padding-top: 3.5rem !important; padding-bottom: 1rem !important; }
    .stMultiSelect [data-baseweb="select"] { min-height: 32px !important; padding-top: 0px !important; padding-bottom: 0px !important; }
    .stMultiSelect [data-baseweb="tag"] { margin-top: 2px !important; margin-bottom: 2px !important; padding-top: 0px !important; padding-bottom: 0px !important; font-size: 13px !important; }
    .stToggle { margin-top: 25px !important; }
    .cacodemon-wrapper { display: none !important; }
    
    div[data-testid="element-container"]:has(.cacodemon-wrapper) + div[data-testid="element-container"] div.stButton button {
        background-image: url('https://doomwiki.org/w/images/c/c3/Cacodemon.png') !important;
        background-size: contain !important;
        background-repeat: no-repeat !important;
        background-position: center !important;
        background-color: transparent !important;
        border: none !important;
        box-shadow: none !important;
        height: 64px !important;
        width: 64px !important;
        image-rendering: pixelated !important;
        transition: transform 0.1s;
    }
    div[data-testid="element-container"]:has(.cacodemon-wrapper) + div[data-testid="element-container"] div.stButton button:active { transform: scale(0.85) !important; }
    div[data-testid="element-container"]:has(.cacodemon-wrapper) + div[data-testid="element-container"] div.stButton button * { display: none !important; }
    
    .doom-container {
        animation: slideDown 1.5s cubic-bezier(0.25, 1, 0.5, 1) forwards;
        border: 4px solid #aa0000;
        border-radius: 8px;
        box-shadow: 0px 10px 30px rgba(255, 0, 0, 0.4);
        margin-top: 15px; margin-bottom: 30px; background-color: #000; padding: 10px;
    }
    @keyframes slideDown { from { opacity: 0; transform: translateY(-30px); } to { opacity: 1; transform: translateY(0); } }
    </style>
""", unsafe_allow_html=True)

# ==========================================
#        BECKHOFF GYÁRI ALKATRÉSZ ADATBÁZIS
# ==========================================
BECKHOFF_DB = {
    # Kuplungok
    "EK1100 (EtherCAT buszkuplung)": {"ebus": 2000, "power": 10.0},
    "EK1101 (EtherCAT buszkuplung ID kapcsolóval)": {"ebus": 2000, "power": 10.0},
    "EK1110 (EtherCAT kiterjesztés végtag sínhez)": {"ebus": -130, "power": 0.0},
    "EK1122 (2 portos EtherCAT elágazás)": {"ebus": -130, "power": 0.0},
    "EK1300 (EtherCAT P buszkuplung)": {"ebus": 2000, "power": 10.0},
    "EK1501 (Száloptikás kuplung)": {"ebus": 2000, "power": 10.0},
    
    # Digitális bemenetek
    "EL1002 (2x DI 24V DC)": {"ebus": -90, "power": 0.0},
    "EL1004 (4x DI 24V DC)": {"ebus": -90, "power": 0.0},
    "EL1008 (8x DI 24V DC)": {"ebus": -90, "power": 0.0},
    "EL1014 (4x DI gyors 10µs)": {"ebus": -90, "power": 0.0},
    "EL1018 (8x DI gyors 10µs)": {"ebus": -90, "power": 0.0},
    "EL1808 (8x DI közös test)": {"ebus": -100, "power": 0.0},
    "EL1809 (16x DI nagy sűrűségű)": {"ebus": -100, "power": 0.0},
    "EL1819 (16x DI gyors 10µs)": {"ebus": -100, "power": 0.0},
    
    # TwinSAFE
    "EL1904 (4x DI 24V DC TwinSAFE)": {"ebus": -200, "power": 0.0},
    "EL1918 (8x DI 24V DC TwinSAFE Logic)": {"ebus": -165, "power": 0.0},
    "EL2904 (4x DO 24V DC TwinSAFE 0.5A)": {"ebus": -221, "power": -2.0},
    
    # Digitális kimenetek
    "EL2002 (2x DO 24V DC 0.5A)": {"ebus": -100, "power": -0.5},
    "EL2004 (4x DO 24V DC 0.5A)": {"ebus": -100, "power": -1.0},
    "EL2008 (8x DO 24V DC 0.5A)": {"ebus": -110, "power": -2.0},
    "EL2024 (4x DO 24V DC 2.0A)": {"ebus": -140, "power": -4.0},
    "EL2809 (16x DO 24V DC 0.5A)": {"ebus": -140, "power": -2.0},
    "EL2819 (16x DO diagnosztikával)": {"ebus": -140, "power": -2.0},
    "EL2624 (4x Relés kimenet)": {"ebus": -120, "power": 0.0},
    
    # Analóg bemenetek
    "EL3002 (2x AI 0-10V 12-bit)": {"ebus": -130, "power": 0.0},
    "EL3004 (4x AI 0-10V 12-bit)": {"ebus": -130, "power": 0.0},
    "EL3008 (8x AI 0-10V 12-bit)": {"ebus": -130, "power": 0.0},
    "EL3024 (4x AI 4-20mA 12-bit)": {"ebus": -130, "power": 0.0},
    "EL3044 (4x AI 0-20mA 12-bit)": {"ebus": -130, "power": 0.0},
    "EL3064 (4x AI 0-10V single-ended)": {"ebus": -130, "power": 0.0},
    "EL3102 (2x AI ±10V 16-bit)": {"ebus": -170, "power": 0.0},
    "EL3104 (4x AI ±10V 16-bit)": {"ebus": -170, "power": 0.0},
    "EL3202 (2x AI PT100)": {"ebus": -190, "power": 0.0},
    "EL3204 (4x AI PT100)": {"ebus": -190, "power": 0.0},
    "EL3208 (8x AI PT100)": {"ebus": -190, "power": 0.0},
    "EL3312 (2x AI Hőelem)": {"ebus": -180, "power": 0.0},
    "EL3314 (4x AI Hőelem)": {"ebus": -180, "power": 0.0},
    
    # Analóg kimenetek
    "EL4002 (2x AO 0-10V 12-bit)": {"ebus": -170, "power": 0.0},
    "EL4004 (4x AO 0-10V 12-bit)": {"ebus": -170, "power": 0.0},
    "EL4008 (8x AO 0-10V 12-bit)": {"ebus": -210, "power": 0.0},
    "EL4024 (4x AO 4-20mA 12-bit)": {"ebus": -170, "power": 0.0},
    "EL4102 (2x AO 0-10V 16-bit)": {"ebus": -250, "power": 0.0},
    "EL4132 (2x AO ±10V 16-bit)": {"ebus": -250, "power": 0.0},
    
    # Rendszer és Speciális Modulok
    "EL5001 (SSI jeladó)": {"ebus": -130, "power": 0.0},
    "EL5101 (Inkrementális jeladó TTL)": {"ebus": -130, "power": 0.0},
    "EL6001 (RS232 modul)": {"ebus": -120, "power": 0.0},
    "EL6002 (2x RS232 modul)": {"ebus": -170, "power": 0.0},
    "EL6070 (EtherCAT licenckulcs terminál)": {"ebus": -140, "power": 0.0},
    "EL6224 (IO-Link Master)": {"ebus": -200, "power": 0.0},
    "EL6601 (Ethernet switch)": {"ebus": -280, "power": 0.0},
    
    # Tápkártyák
    "EL9011 (Buszlezáró - End Cap)": {"ebus": 0, "power": 0.0},
    "EL9100 (24V Power érintkező táp)": {"ebus": 0, "power": 10.0},
    "EL9110 (24V Power diagnosztikával)": {"ebus": 0, "power": 10.0},
    "EL9200 (24V Power biztosítékkal)": {"ebus": 0, "power": 10.0},
    "EL9410 (E-Bus tápfrissítő & Power táp)": {"ebus": 2000, "power": 10.0},
    "EL9505 (5V DC belső táp)": {"ebus": -10, "power": 0.0},
    "EL9510 (10V DC belső táp)": {"ebus": -10, "power": 0.0}
}

# --- ADATBÁZIS KEZELŐ (JSON) ---
DB_FILE = 'db_users.json'

def load_db():
    if not os.path.exists(DB_FILE):
        default_db = {
            "admin": {"password": "admin", "todo": ["Üdv!"], "archive": [], "status": "Aktív"},
            "custom_cards": {},
            "custom_zones": ["E-Bus Fő", "Fő betáp (24V)"]
        }
        with open(DB_FILE, 'w', encoding='utf-8') as f: json.dump(default_db, f, ensure_ascii=False, indent=4)
        return default_db
    with open(DB_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
        if "custom_cards" not in data: data["custom_cards"] = {}
        if "custom_zones" not in data: data["custom_zones"] = ["E-Bus Fő", "Fő betáp (24V)"]
        save_db(data)
        return data

def save_db(data):
    with open(DB_FILE, 'w', encoding='utf-8') as f: json.dump(data, f, ensure_ascii=False, indent=4)

db = load_db()

if 'user' not in st.session_state: st.session_state.user = None
if 'guest_todo' not in st.session_state: st.session_state.guest_todo = []
if 'editing_index' not in st.session_state: st.session_state.editing_index = None
if 'doom_clicks' not in st.session_state: st.session_state.doom_clicks = 0

# --- MODUL ÉS ZÓNA ADATBÁZIS ÖSSZEVONÁS ---
if 'custom_cards' not in st.session_state: st.session_state.custom_cards = db.get("custom_cards", {})
if 'custom_zones' not in st.session_state: st.session_state.custom_zones = db.get("custom_zones", ["E-Bus Fő", "Fő betáp (24V)"])

FULL_BECKHOFF_DB = BECKHOFF_DB.copy()
FULL_BECKHOFF_DB.update(st.session_state.custom_cards)

# --- SORSZÁMOS ALAPKONFIGURÁCIÓ (Csak az EK1100) ---
if 'beckhoff_rack_simple' not in st.session_state or "Sorszám" not in st.session_state.beckhoff_rack_simple.columns:
    st.session_state.beckhoff_rack_simple = pd.DataFrame([
        {"Sorszám": 10, "Terminál típus": "EK1100 (EtherCAT buszkuplung)", "Darabszám": 1, "E-Bus Zóna": "E-Bus Fő", "Terepi Zóna": "Fő betáp (24V)"}
    ])

def save_todo_logic(new_list):
    if st.session_state.user:
        db[st.session_state.user]["todo"] = new_list
        save_db(db)
    else: st.session_state.guest_todo = new_list

current_todo_list = db[st.session_state.user]["todo"] if st.session_state.user else st.session_state.guest_todo

# --- 📝 OLDALSÁV ---
with st.sidebar:
    if st.session_state.user is None:
        st.title("👤 Fiókkezelés")
        st.info("Vendég mód: A munkád és jegyzeteid nem mentődnek tartósan.")
        with st.expander("🔑 Bejelentkezés / Belépés", expanded=True):
            login_user = st.text_input("Felhasználónév")
            login_pass = st.text_input("Jelszó", type="password")
            if st.button("Belépés", use_container_width=True):
                if login_user in db and db[login_user]["password"] == login_pass:
                    if db[login_user].get("status", "Aktív") == "Letiltva": st.error("❌ Ez a fiók le van tiltva!")
                    else:
                        st.session_state.user = login_user
                        st.success("Sikeres belépés!")
                        st.rerun()
                else: st.error("Hibás adatok!")
    else:
        col_caco, col_text = st.columns([1, 3])
        with col_caco:
            st.markdown('<div class="cacodemon-wrapper"></div>', unsafe_allow_html=True)
            if st.button("👾", key="caco_btn", help="Kattints rám gyorsan 6-szor..."): st.session_state.doom_clicks += 1
        with col_text: st.markdown("<h2 style='margin-top: -10px;'>Fiókkezelés</h2>", unsafe_allow_html=True)
            
        st.success(f"Bejelentkezve: **{st.session_state.user}**")
        
        if st.session_state.user == "admin":
            with st.expander("👥 Felhasználók kezelése (Admin)", expanded=False):
                add_user = st.text_input("Új felhasználónév", key="admin_add_user").strip()
                add_pass = st.text_input("Új jelszó", type="password", key="admin_add_pass").strip()
                if st.button("Fiók létrehozása", use_container_width=True):
                    if add_user and add_pass:
                        if add_user in db: st.error("❌ Már létezik!")
                        else:
                            db[add_user] = {"password": add_pass, "todo": [], "archive": [], "status": "Aktív"}
                            save_db(db)
                            st.success(f"✅ '{add_user}' regisztrálva!")
                            st.rerun()
                    else: st.error("❌ A mezők nem lehetnek üresek!")
                st.divider()
                user_list = [u for u in db.keys() if u != "admin" and u not in ["custom_cards", "custom_zones"]]
                if user_list:
                    manage_user = st.selectbox("Szerkesztendő fiók", ["-- Válassz --"] + user_list)
                    if manage_user != "-- Válassz --":
                        current_status = db[manage_user].get("status", "Aktív")
                        st.write(f"Állapot: **{current_status}**")
                        mod_p = st.text_input("Új jelszó kényszerítése", type="password", key=f"admin_mod_p_{manage_user}").strip()
                        if st.button("🔒 Jelszó csere", key=f"admin_btn_p_{manage_user}", use_container_width=True):
                            if mod_p:
                                db[manage_user]["password"] = mod_p
                                save_db(db)
                                st.rerun()
                        if current_status == "Aktív":
                            if st.button("🚫 Letiltás", key=f"admin_btn_lock_{manage_user}", use_container_width=True):
                                db[manage_user]["status"] = "Letiltva"
                                save_db(db); st.rerun()
                        else:
                            if st.button("✅ Aktiválás", key=f"admin_btn_unlock_{manage_user}", use_container_width=True):
                                db[manage_user]["status"] = "Aktív"
                                save_db(db); st.rerun()
                        if st.button("🗑️ Törlés", key=f"admin_btn_del_{manage_user}", use_container_width=True):
                            del db[manage_user]
                            save_db(db); st.rerun()
        
        with st.expander("⚙️ Saját jelszó módosítása"):
            old_p = st.text_input("Jelenlegi jelszó", type="password")
            new_p1 = st.text_input("Új jelszó", type="password")
            new_p2 = st.text_input("Új jelszó megerősítése", type="password")
            if st.button("Mentés", use_container_width=True):
                if old_p == db[st.session_state.user]["password"] and new_p1 == new_p2 and new_p1:
                    db[st.session_state.user]["password"] = new_p1
                    save_db(db); st.success("✅ Frissítve!")
                else: st.error("❌ Hibás adatok!")
        
        if st.button("🚪 Kijelentkezés", use_container_width=True):
            st.session_state.user = None
            st.session_state.doom_clicks = 0
            st.rerun()
            
    st.divider()
    st.title("📝 Napi Teendők")
    if st.session_state.user: st.caption("☁️ Szinkronizálva a fiókoddal.")
    else: st.caption("⚠️ Vendégként frissítéskor eltűnnek.")
        
    col1, col2 = st.columns([3, 1])
    with col1: uj_feladat = st.text_input("Új feladat:", label_visibility="collapsed", placeholder="Új feladat...", key="uj_feladat_input")
    with col2:
        if st.button("➕", use_container_width=True, key="add_button"):
            if uj_feladat.strip():
                current_todo_list.append(uj_feladat.strip())
                save_todo_logic(current_todo_list); st.rerun()
    st.write("")
    
    if not current_todo_list: st.caption("Nincsenek teendők.")
    else:
        for i, feladat in enumerate(current_todo_list):
            if st.session_state.editing_index == i:
                c1, c2 = st.columns([3, 1])
                with c1: edit_val = st.text_input(f"edit_{i}", value=feladat, label_visibility="collapsed")
                with c2:
                    if st.button("✅", key=f"save_{i}", use_container_width=True):
                        current_todo_list[i] = edit_val
                        st.session_state.editing_index = None
                        save_todo_logic(current_todo_list); st.rerun()
            else:
                c1, c2, c3 = st.columns([5, 2, 2])
                with c1: st.markdown(f"<div style='margin-top: 6px;'>• {feladat}</div>", unsafe_allow_html=True)
                with c2:
                    if st.button("✏️", key=f"edit_{i}", use_container_width=True):
                        st.session_state.editing_index = i; st.rerun()
                with c3:
                    if st.button("❌", key=f"del_{i}", use_container_width=True):
                        current_todo_list.pop(i)
                        save_todo_logic(current_todo_list); st.rerun()

    if st.session_state.user:
        st.divider()
        st.title("🗃️ Munkák Archívuma")
        user_archive = db[st.session_state.user].get("archive", [])
        if not user_archive: st.caption("Nincs elmentett projekt.")
        else:
            for arch in reversed(user_archive):
                with st.expander(f"📁 {arch['date']} - {arch['project']}"):
                    st.write(f"**Állomás:** {arch['station']}")
                    st.write(f"**Tételek:** {arch['row_count']} db")
                    arch_df = pd.DataFrame(arch['data'])
                    arch_buffer = io.BytesIO()
                    with pd.ExcelWriter(arch_buffer, engine='openpyxl') as writer: arch_df.to_excel(writer, index=False, sheet_name='Archivalt')
                    
                    c_dl, c_del = st.columns(2)
                    with c_dl: st.download_button(label="📥 Letöltés", data=arch_buffer.getvalue(), file_name=f"Archivum_{arch['project']}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_{arch['id']}")
                    with c_del:
                        if st.button("🗑️ Törlés", key=f"del_arch_{arch['id']}"):
                            db[st.session_state.user]["archive"] = [a for a in db[st.session_state.user]["archive"] if a["id"] != arch["id"]]
                            save_db(db); st.rerun()

# --- SEGÉDFÜGGVÉNYEK ---
def get_excel_modified_date(uploaded_file):
    try:
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, read_only=True)
        mod_time = wb.properties.modified
        uploaded_file.seek(0)
        if mod_time: return mod_time.strftime("%Y. %m. %d. %H:%M")
        return "Ismeretlen dátum"
    except Exception:
        uploaded_file.seek(0)
        return "Ismeretlen dátum"

def get_column_safe(df, col_index):
    if col_index < len(df.columns):
        s = df.iloc[:, col_index].astype(str).str.strip()
        return s.replace(['nan', 'None', '', 'NaN', '<NA>'], pd.NA)
    return pd.Series([pd.NA] * len(df), index=df.index)

# --- FŐ FELÜLET FEJLÉC ---
col_title, col_info = st.columns([3, 1])
with col_title: st.title("⚙️ Ekszi segéd")
with col_info: info_placeholder = st.empty()

# --- 🎮 DOOM TETŐ-ABLAK EASTER EGG ---
if st.session_state.get('doom_clicks', 0) >= 6:
    st.error("🩸 **Kód aktiválva. IDDQD! RIP AND TEAR! Kattints a középső kerek 'Bekapcsoló' gombra a játékhoz!**")
    st.markdown('<div class="doom-container">', unsafe_allow_html=True)
    components.html(
        """<div style="display: flex; justify-content: center; align-items: center; background-color: #000; width: 100%; height: 100%;"><iframe src="https://archive.org/embed/doom_dos" width="800" height="600" frameborder="0" webkitallowfullscreen="true" mozallowfullscreen="true" allowfullscreen></iframe></div>""", height=620,
    )
    st.markdown('</div>', unsafe_allow_html=True)

# --- FÜLES RENDSZER (TABS) ---
tab1, tab2 = st.tabs(["📊 Alkatrészkezelő", "⚡ Beckhoff modul kalkulátor"])

# ------------------------------------------
#   1. TAB: ALKATRÉSZKEZELŐ
# ------------------------------------------
with tab1:
    st.write("Töltsd fel az EPLAN-ból kimentett Excel fájlt, szerkeszd a tételeket, majd töltsd le az eredményt!")
    st.subheader("📁 Fájlok feltöltése")
    col1, col2 = st.columns(2)

    with col1:
        eplan_ph = st.empty()
        eplan_file = st.file_uploader("eplan_upload", type=["xlsx", "xls"], label_visibility="collapsed")
        if eplan_file is not None:
            eplan_date = get_excel_modified_date(eplan_file)
            eplan_ph.markdown(f"<div style='display: flex; justify-content: space-between; align-items: center; padding-bottom: 5px;'><span style='font-size: 14px;'>1. EPLAN Excel feltöltése</span><span style='color: #888; font-size: 12px;'>🕒 <b>Utolsó mentés:</b> {eplan_date}</span></div>", unsafe_allow_html=True)
        else: eplan_ph.markdown("<div style='font-size: 14px; padding-bottom: 5px;'>1. EPLAN Excel feltöltése</div>", unsafe_allow_html=True)
        
    with col2:
        raktar_ph = st.empty()
        raktar_file = st.file_uploader("raktar_upload", type=["xlsx", "xls"], label_visibility="collapsed")
        if raktar_file is not None:
            raktar_date = get_excel_modified_date(raktar_file)
            raktar_ph.markdown(f"<div style='display: flex; justify-content: space-between; align-items: center; padding-bottom: 5px;'><span style='font-size: 14px;'>2. DM aktuális készlet</span><span style='color: #888; font-size: 12px;'>🕒 <b>Utolsó mentés:</b> {raktar_date}</span></div>", unsafe_allow_html=True)
        else: raktar_ph.markdown("<div style='font-size: 14px; padding-bottom: 5px;'>2. DM aktuális készlet (Opcionális)</div>", unsafe_allow_html=True)

    if eplan_file is not None:
        projekt_szam, allomas_szam = "Nincs megadva", "Nincs megadva"
        try:
            df_check = pd.read_excel(eplan_file, nrows=5)
            is_resumed = "Beszerzés státusza" in df_check.columns

            if is_resumed:
                df_eplan = pd.read_excel(eplan_file).fillna("")
                info_placeholder.markdown(f'<div style="background-color: #f0f2f6; padding: 15px; border-radius: 10px; border-left: 5px solid #28a745;"><div style="font-size: 14px; color: #555;">🔄 <b>Állapot:</b> Folytatott munka betöltve</div></div>', unsafe_allow_html=True)
            else:
                df_eplan_raw = pd.read_excel(eplan_file, header=None)
                nyers_projekt = str(df_eplan_raw.iloc[0, 1]).strip() if pd.notna(df_eplan_raw.iloc[0, 1]) else "Nincs megadva"
                if nyers_projekt != "Nincs megadva":
                    projekt_szam = nyers_projekt[:-5] if len(nyers_projekt) > 5 else nyers_projekt
                    if len(projekt_szam) > 11: projekt_szam = projekt_szam[:4] + '/' + projekt_szam[5:11] + '-' + projekt_szam[12:]
                    elif len(projekt_szam) > 4: projekt_szam = projekt_szam[:4] + '/' + projekt_szam[5:]
                
                allomas_szam = str(df_eplan_raw.iloc[1, 1]).strip() if pd.notna(df_eplan_raw.iloc[1, 1]) else "Nincs megadva"
                info_placeholder.markdown(f'<div style="background-color: #f0f2f6; padding: 15px; border-radius: 10px; border-left: 5px solid #005A9C;"><div style="font-size: 14px; color: #555;">📁 <b>Projekt:</b> {projekt_szam}</div><div style="font-size: 14px; color: #555; margin-top: 5px;">📍 <b>Állomás:</b> {allomas_szam}</div></div>', unsafe_allow_html=True)
                
                df_eplan = df_eplan_raw.iloc[5:].reset_index(drop=True)
                df_eplan.columns = df_eplan_raw.iloc[4]
                df_eplan.dropna(how='all', inplace=True)
                df_eplan = df_eplan.reset_index(drop=True)
                
                eplan_a_col_name = df_eplan.columns[0] if len(df_eplan.columns) > 0 else None
                eplan_c_series = get_column_safe(df_eplan, 2) 
                eplan_d_series = get_column_safe(df_eplan, 3) 
                eplan_qty_series = get_column_safe(df_eplan, 4) 
                
                valid_eplan_mask = eplan_d_series.fillna(eplan_c_series).notna() & (eplan_d_series.fillna(eplan_c_series) != "")
                
                df_eplan = df_eplan[valid_eplan_mask].reset_index(drop=True)
                eplan_c_series, eplan_d_series, eplan_qty_series = eplan_c_series[valid_eplan_mask].reset_index(drop=True), eplan_d_series[valid_eplan_mask].reset_index(drop=True), eplan_qty_series[valid_eplan_mask].reset_index(drop=True)
                
                if eplan_a_col_name and eplan_a_col_name in df_eplan.columns: df_eplan = df_eplan.drop(columns=[eplan_a_col_name])
                df_eplan = df_eplan.fillna("")
                df_eplan.insert(0, "Beszerzés státusza", "Kiválasztandó...")
        except Exception as e: st.error(f"❌ Hiba az EPLAN fájl beolvasásakor! Részletek: {e}"); st.stop()
        
        if raktar_file is not None:
            try:
                df_raktar_raw = pd.read_excel(raktar_file, header=None)
                kulcsszavak = {'cikkszam': ['cikkszám', 'típus', 'cikk', 'megrendelési szám', 'anyagszám', 'material', 'azonosító'], 'mennyiseg': ['mennyiség', 'készlet', 'db', 'stock', 'szabad', 'raktárkészlet', 'menny.'], 'sarzs': ['raktárhely', 'hely', 'fiók', 'polc', 'bin', 'storage', 'tárhely', 'sarzs'], 'raktar': ['raktár', 'wh', 'warehouse', 'rkg']}
                fejlec_sor_idx, c_idx, q_idx, h_idx, r_idx = -1, -1, -1, -1, -1

                for row_idx, row in df_raktar_raw.head(20).iterrows():
                    temp_c, temp_q, temp_h, temp_r = -1, -1, -1, -1
                    for col_idx, cell_value in enumerate(row):
                        cell_str = str(cell_value).lower().strip()
                        if temp_c == -1 and any(k in cell_str for k in kulcsszavak['cikkszam']): temp_c = col_idx
                        elif temp_q == -1 and any(k in cell_str for k in kulcsszavak['mennyiseg']): temp_q = col_idx
                        elif temp_h == -1 and any(k in cell_str for k in kulcsszavak['sarzs']): temp_h = col_idx
                        elif temp_r == -1 and any(k in cell_str for k in kulcsszavak['raktar']): temp_r = col_idx
                    if temp_c != -1 and temp_q != -1:
                        fejlec_sor_idx, c_idx, q_idx, h_idx, r_idx = row_idx, temp_c, temp_q, temp_h, temp_r
                        break
                
                if fejlec_sor_idx != -1:
                    if h_idx == -1: h_idx = 0 
                    actual_helyek = df_raktar_raw.iloc[:, h_idx].astype(str).str.strip()
                    mask_empty = actual_helyek.str.lower().isin(['nan', 'none', 'null', '<na>', '', '0', '0.0'])
                    if c_idx == h_idx:
                        qty_col = df_raktar_raw.iloc[:, q_idx].astype(str).str.lower().str.strip()
                        actual_helyek[~((~mask_empty) & qty_col.isin(['nan', 'none', 'null', '<na>', '', '0', '0.0']))] = None
                    else: actual_helyek[mask_empty] = None
                    actual_helyek.iloc[fejlec_sor_idx] = None
                    actual_helyek = actual_helyek.ffill().fillna("Nincs megadva")
                    
                    if r_idx != -1:
                        actual_raktarak = df_raktar_raw.iloc[:, r_idx].astype(str).str.strip()
                        actual_raktarak[actual_raktarak.str.lower().isin(['nan', 'none', 'null', '<na>', '', '0', '0.0'])] = None
                        actual_raktarak.iloc[fejlec_sor_idx] = None
                        actual_raktarak = actual_raktarak.ffill().fillna("Ismeretlen Raktár")
                    else: actual_raktarak = pd.Series(["Nincs Raktár Oszlop"] * len(df_raktar_raw))
                    
                    df_raktar = df_raktar_raw.iloc[fejlec_sor_idx + 1:].reset_index(drop=True)
                    raktar_f_helyek = actual_helyek.iloc[fejlec_sor_idx + 1:].reset_index(drop=True)
                    raktar_f_nevek = actual_raktarak.iloc[fejlec_sor_idx + 1:].reset_index(drop=True)
                    
                    raktar_vegleges_cikkszam = df_raktar.iloc[:, c_idx].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).replace(['nan', 'None', '', 'NaN', '<NA>'], "-")
                    tiszta_raktar_qty = pd.to_numeric(df_raktar.iloc[:, q_idx].astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)
                    
                    df_raktar_clean = pd.DataFrame({'Cikkszám': raktar_vegleges_cikkszam, 'Mennyiség': tiszta_raktar_qty, 'Hely': raktar_f_helyek, 'Raktár_Neve': raktar_f_nevek})
                    df_raktar_clean = df_raktar_clean[df_raktar_clean['Cikkszám'] != "-"]
                    
                    if st.session_state.get('ignore_projektek', False): df_raktar_clean = df_raktar_clean[~df_raktar_clean['Raktár_Neve'].astype(str).str.contains('Projektek', case=False, na=False)]
                    
                    df_shelf_totals = df_raktar_clean.groupby(['Cikkszám', 'Hely'], as_index=False)['Mennyiség'].sum()
                    def format_shelf_qty(v): return str(int(v)) if v == int(v) else str(v)
                    df_shelf_totals['Hely_Kombinalt'] = df_shelf_totals['Hely'].astype(str) + " (" + df_shelf_totals['Mennyiség'].apply(format_shelf_qty) + " db)"
                    
                    aggregated_raktar = df_shelf_totals.groupby('Cikkszám').agg({'Mennyiség': 'sum', 'Hely_Kombinalt': lambda x: ", ".join(x)})
                    raktar_hely_dict, raktar_qty_dict = aggregated_raktar['Hely_Kombinalt'].to_dict(), aggregated_raktar['Mennyiség'].to_dict()
                else: st.error("❌ Nem találtam oszlopokat a raktári fájlban!"); st.stop()

                if is_resumed:
                    cikkszam_col = next((col for col in df_eplan.columns if 'cikkszám' in str(col).lower() or 'típus' in str(col).lower()), df_eplan.columns[3])
                    eplan_vegleges_cikkszam = df_eplan[cikkszam_col].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).replace(['nan', 'None', '', 'NaN'], "-")
                    qty_col = next((col for col in df_eplan.columns if 'mennyiség' in str(col).lower() or 'db' in str(col).lower()), df_eplan.columns[4])
                    tiszta_eplan_qty = pd.to_numeric(df_eplan[qty_col].astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)
                else:
                    eplan_vegleges_cikkszam = eplan_d_series.fillna(eplan_c_series).astype(str).str.strip().str.replace(r'\.0$', '', regex=True).fillna("-")
                    tiszta_eplan_qty = pd.to_numeric(eplan_qty_series.astype(str).str.replace(r'[^\d\.]', '', regex=True), errors='coerce').fillna(0)
                    
                for col in ['Raktáron találva', 'Raktárhely', 'Raktári készlet']:
                    if col in df_eplan.columns: df_eplan = df_eplan.drop(columns=[col])
                
                df_eplan['Raktárhely'] = eplan_vegleges_cikkszam.map(raktar_hely_dict).fillna("-")
                df_eplan['Raktári készlet'] = eplan_vegleges_cikkszam.map(raktar_qty_dict).fillna(0)
                
                def ertekel_mennyiseg(row, idx):
                    if row['Raktárhely'] == "-": return "❌ Nem"
                    e_qty = tiszta_eplan_qty.iloc[idx]
                    r_qty = row['Raktári készlet']
                    if r_qty <= 0: return "❌ Nem"
                    elif r_qty >= e_qty: return "✅ Igen"
                    else: return "⚠️ Részleges"
                
                df_eplan['Raktáron találva'] = [ertekel_mennyiseg(row, idx) for idx, row in df_eplan.iterrows()]
                mask_igen = (df_eplan['Raktáron találva'] == "✅ Igen") & (df_eplan['Beszerzés státusza'] == "Kiválasztandó...")
                df_eplan.loc[mask_igen, "Beszerzés státusza"] = "Áttároltatható"
                mask_nem = (df_eplan['Raktáron találva'].isin(["⚠️ Részleges", "❌ Nem"])) & (df_eplan['Beszerzés státusza'] == "Kiválasztandó...")
                df_eplan.loc[mask_nem, "Beszerzés státusza"] = "Rendelni"
                
                col_keszlet, col_hely, col_talalat = df_eplan.pop('Raktári készlet'), df_eplan.pop('Raktárhely'), df_eplan.pop('Raktáron találva')
                df_eplan.insert(0, 'Raktáron találva', col_talalat)
                df_eplan.insert(1, 'Raktárhely', col_hely)
                df_eplan.insert(2, 'Raktári készlet', col_keszlet)
                st.success("✅ Raktárkészlet és Sarzs kódok sikeresen beolvasva!")
            except Exception as e: st.error(f"❌ Hiba történt a raktár beolvasásakor: {e}")

        # --- TÁBLÁZAT ÉS SZŰRŐK MEGJELENÍTÉSE ---
        st.divider()
        st.subheader("📝 Adatok szerkesztése és szűrése")
        
        df_to_display = df_eplan.copy()
        
        if 'Raktáron találva' in df_to_display.columns:
            col_filters, col_toggle = st.columns([2, 1])
            with col_filters:
                st.write("🔍 **Szűrés raktári találat alapján:**")
                f1, f2, f3 = st.columns(3)
                with f1: show_igen = st.checkbox("✅ Igen", value=True)
                with f2: show_reszleges = st.checkbox("⚠️ Részleges", value=True)
                with f3: show_nem = st.checkbox("❌ Nem", value=True)
                
                kivalasztott_szurok = []
                if show_igen: kivalasztott_szurok.append("✅ Igen")
                if show_reszleges: kivalasztott_szurok.append("⚠️ Részleges")
                if show_nem: kivalasztott_szurok.append("❌ Nem")

            with col_toggle:
                st.write("⚙️ **Készlet szabály:**")
                st.toggle("🚫 'Projektek' raktárak elrejtése", key="ignore_projektek")
                
            df_to_display = df_to_display[df_to_display['Raktáron találva'].isin(kivalasztott_szurok)]

            with st.expander("🔍 Gyorskeresés a táblázatban", expanded=False):
                c_label, c_input = st.columns([1, 4])
                with c_label: st.markdown("<div style='margin-top: 5px; font-weight: bold;'>Keresőszó:</div>", unsafe_allow_html=True)
                with c_input: search_text = st.text_input("Keresés", label_visibility="collapsed", placeholder="Cikkszám, Megnevezés, Gyártó, stb...")
            
            if search_text:
                mask = df_to_display.apply(lambda row: row.astype(str).str.contains(search_text, case=False, na=False).any(), axis=1)
                df_to_display = df_to_display[mask]

        def color_rows(row):
            if 'Raktáron találva' in row and row['Raktáron találva'] in ["✅ Igen", "⚠️ Részleges"]:
                hely = str(row.get('Raktárhely', '')).strip()
                if "Projektek" not in hely: return ['background-color: #C6E0B4; color: black;'] * len(row) 
                else: return ['background-color: #FFF2CC; color: black;'] * len(row) 
            return [''] * len(row)

        styled_eplan = df_to_display.style.apply(color_rows, axis=1)

        edited_df = st.data_editor(
            styled_eplan,
            column_config={
                "Beszerzés státusza": st.column_config.SelectboxColumn(
                    "Beszerzés státusza",
                    options=["Kiválasztandó...", "Rendelni", "Raktárból", "Áttároltatható", "Közös rendelés", "Már megrendelve", "Nem kell"],
                    required=True
                )
            }, use_container_width=True, hide_index=True, num_rows="dynamic"
        )

        st.divider()
        st.subheader("📥 Exportálás és Mentés")
        col_exp1, col_exp2, col_exp3 = st.columns([1, 1, 1])
        
        with col_exp1:
            st.info("💡 **Lokális Mentés:** (Folytatáshoz)")
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer: edited_df.to_excel(writer, index=False, sheet_name='Feldolgozott_Lista')
            st.download_button(label="💾 Mentés (Excel)", data=buffer.getvalue(), file_name="EPLAN_munka_mentese.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        with col_exp2:
            st.info("💡 **Csillagpont:** Ajánlatkérő")
            if st.button("📤 Csillagpont generálás"):
                try:
                    wb = openpyxl.load_workbook("sablon.xlsx")
                    ws = wb["VM_BOM"]
                    ws["B1"], ws["B2"] = projekt_szam, allomas_szam
                    csillagpont_df = edited_df[edited_df["FORGALMAZÓ"].astype(str).str.contains("Csillagpont", case=False, na=False)]
                    if csillagpont_df.empty: st.warning("⚠️ Nincs Csillagpontos tétel.")
                    else:
                        for r_idx, (index, row) in enumerate(csillagpont_df.iterrows(), start=6):
                            ws.cell(row=r_idx, column=1, value=row["MEGNEVEZÉS 1"]); ws.cell(row=r_idx, column=2, value=row["MEGRENDELÉSI SZÁM"])
                            ws.cell(row=r_idx, column=3, value=row["MENNYISÉG"]); ws.cell(row=r_idx, column=4, value=row["MÉRTÉKEGYSÉG"])
                            ws.cell(row=r_idx, column=5, value=row["GYÁRTÓ"])
                        output = io.BytesIO()
                        wb.save(output); output.seek(0)
                        st.download_button(label="📥 Letöltés: Csillagpont", data=output.getvalue(), file_name="Csillagpont_ajanlatkeres.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e: st.error(f"❌ Hiba: {e}")

        with col_exp3:
            if st.session_state.user:
                st.info("💡 **Archiválás:** Mentés a fiókodba")
                if st.button("🗃️ Mentés az Archívumba"):
                    arch_id = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                    arch_date = datetime.datetime.now().strftime("%Y. %m. %d. %H:%M")
                    save_data = edited_df.fillna("").to_dict(orient='records')
                    if "archive" not in db[st.session_state.user]: db[st.session_state.user]["archive"] = []
                    db[st.session_state.user]["archive"].append({"id": arch_id, "date": arch_date, "project": projekt_szam, "station": allomas_szam, "row_count": len(edited_df), "data": save_data})
                    save_db(db); st.success("✅ Elmentve az Archívumba!"); st.rerun()
            else: st.warning("🔒 Az Archívum mentéshez be kell jelentkezned a bal oldalon.")
    else:
        st.write("")
        st.info("👋 **Üdvözöllek!** A kezdéshez kérlek töltsd fel az **1. EPLAN Excel** fájlt fent.")

# ------------------------------------------
#   2. TAB: BECKHOFF OKOS KALKULÁTOR
# ------------------------------------------
with tab2:
    st.markdown("### ⚡ Beckhoff modul kalkulátor")
    st.write("A rendszer **KETTÉVÁLASZTVA** számolja a belső E-Bus (5V) és a Terepi (24V) áramköröket!")
    
    with st.expander("⚙️ Kalkulátor Beállítások és Új Kártyák", expanded=False):                    
        st.markdown("**1. Új, egyedi kártyatípus rögzítése az adatbázisba:**")
        c_name = st.text_input("Kártya pontos neve (pl. ELX3162):").strip()
        col_c1, col_c2 = st.columns(2)
        with col_c1: c_ebus = st.number_input("E-Bus áramérték (mA)", value=0, step=10, help="Fogyasztónál NEGATÍV, tápegységnél POZITÍV")
        with col_c2: c_power = st.number_input("Power Érintkező terhelés (A)", value=0.0, step=0.5, help="Terhelésnél NEGATÍV, tápbevezetőnél POZITÍV")
        if st.button("Kártyatípus mentése", use_container_width=True):
            if c_name:
                st.session_state.custom_cards[c_name] = {"ebus": int(c_ebus), "power": float(c_power)}
                db["custom_cards"][c_name] = {"ebus": int(c_ebus), "power": float(c_power)}
                save_db(db); st.success(f"✅ A(z) '{c_name}' mentve!"); st.rerun()
                
        st.divider()
        st.markdown("**2. Automatikus táp beavatkozás:**")
        c_opt1, c_opt2 = st.columns(2)
        with c_opt1: auto_power_card = st.selectbox("Power (terepi) áramhiány esetén:", options=["EL9100 (24V Power érintkező táp)", "EL9410 (E-Bus tápfrissítő & Power táp)", "EL9110 (24V Power diagnosztikával)", "EL9200 (24V Power biztosítékkal)"])
        with c_opt2: auto_ebus_card = st.selectbox("E-Bus áramhiány esetén:", options=["EL9410 (E-Bus tápfrissítő & Power táp)"])

    st.write("🛠️ **Állítsd össze a sín (Rack) konfigurációját:**")
    st.caption("💡 **TIPP:** A Zóna legördülőben válaszd az **'➕ Új zóna megadása...'** opciót egy teljesen új potenciál létrehozásához!")
    
    zone_options = [""] + st.session_state.custom_zones + ["➕ Új zóna megadása..."]
    
    edited_rack = st.data_editor(
        st.session_state.beckhoff_rack_simple,
        column_config={
            "Sorszám": st.column_config.NumberColumn("Sorszám", min_value=1, step=1, help="Rendezd át a kártyákat a sorszám módosításával!"),
            "Terminál típus": st.column_config.SelectboxColumn("Kártya Típusa", options=list(FULL_BECKHOFF_DB.keys()), required=True),
            "Darabszám": st.column_config.NumberColumn("Darabszám", min_value=1, default=1, step=1),
            "E-Bus Zóna": st.column_config.SelectboxColumn("E-Bus Zóna", options=zone_options),
            "Terepi Zóna": st.column_config.SelectboxColumn("Terepi Zóna", options=zone_options)
        }, num_rows="dynamic", use_container_width=True, hide_index=True, key="beckhoff_auto_editor"
    )
    
    # ⚡ AUTOMATIKUS SORSZÁMOZÁS LOGIKA (Ha valaki hozzáad egy új sort)
    mask_empty = edited_rack["Sorszám"].isna() | (edited_rack["Sorszám"] == 0)
    if mask_empty.any():
        max_sorszam = edited_rack.loc[~mask_empty, "Sorszám"].max()
        if pd.isna(max_sorszam): max_sorszam = 0
        for i in edited_rack[mask_empty].index:
            max_sorszam += 10
            edited_rack.at[i, "Sorszám"] = max_sorszam
        st.session_state.beckhoff_rack_simple = edited_rack
        st.rerun()

    # ⚡ ÚJ ZÓNA FELVÉTELE A TÁBLÁZATBÓL
    mask_ebus = edited_rack["E-Bus Zóna"] == "➕ Új zóna megadása..."
    mask_power = edited_rack["Terepi Zóna"] == "➕ Új zóna megadása..."
    
    if mask_ebus.any() or mask_power.any():
        st.warning("💡 Kiválasztottad az 'Új zóna megadása...' opciót. Kérlek írd be az új zóna nevét!")
        c_uj1, c_uj2 = st.columns([3,1])
        with c_uj1:
            uj_zona_nev = st.text_input("Az új zóna neve (pl. Biztonsági kör 24V):", key="uj_zona_input")
        with c_uj2:
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            if st.button("Hozzáadás a listához", use_container_width=True):
                if uj_zona_nev.strip():
                    uj_zona_nev = uj_zona_nev.strip()
                    if uj_zona_nev not in st.session_state.custom_zones:
                        st.session_state.custom_zones.append(uj_zona_nev)
                        db["custom_zones"] = st.session_state.custom_zones
                        save_db(db)
                    
                    edited_rack.loc[mask_ebus, "E-Bus Zóna"] = uj_zona_nev
                    edited_rack.loc[mask_power, "Terepi Zóna"] = uj_zona_nev
                    st.session_state.beckhoff_rack_simple = edited_rack
                    st.rerun()
                else:
                    st.error("A név nem lehet üres!")
    else:
        # CSAK AKKOR FUT LE A SZIMULÁCIÓ HA NINCS FÜGGŐBEN LÉVŐ ÚJ ZÓNA
        st.divider()
        st.markdown("#### 🎯 Automatikusan Optimalizált Kiosztás (Sorszám alapján rendezve!)")

        # 1. ELŐSZÖR SORBA RENDEZZÜK A FELHASZNÁLÓI SORSZÁM ALAPJÁN
        processed_rack = edited_rack.sort_values(by="Sorszám").reset_index(drop=True)

        final_rack = []
        ebus_balance = 0
        power_balance = 0
        ebus_zone = "E-Bus Fő"
        power_zone = "Fő betáp (24V)"
        real_pos = 1
        auto_added_count = 0

        for idx, row in processed_rack.iterrows():
            t_type = row.get("Terminál típus", "")
            qty = int(row.get("Darabszám", 1))
            req_ebus = str(row.get("E-Bus Zóna", "")).replace("nan", "").strip()
            req_power = str(row.get("Terepi Zóna", "")).replace("nan", "").strip()

            if not t_type or t_type not in FULL_BECKHOFF_DB or qty < 1:
                continue

            card_ebus = FULL_BECKHOFF_DB[t_type]["ebus"]
            card_power = FULL_BECKHOFF_DB[t_type]["power"]
            
            is_ebus_refresher = ("EK" in t_type) or ("EL94" in t_type) or (t_type in st.session_state.custom_cards and card_ebus > 0)
            is_power_refresher = ("EK" in t_type) or ("EL91" in t_type) or ("EL92" in t_type) or ("EL94" in t_type) or (t_type in st.session_state.custom_cards and card_power > 0)

            for i in range(qty):
                curr_req_ebus = req_ebus if i == 0 else ""
                curr_req_power = req_power if i == 0 else ""

                if is_ebus_refresher:
                    ebus_balance = card_ebus
                    if curr_req_ebus: ebus_zone = curr_req_ebus
                if is_power_refresher:
                    power_balance = card_power
                    if curr_req_power: power_zone = curr_req_power

                if curr_req_ebus and curr_req_ebus != ebus_zone and not is_ebus_refresher:
                    ebus_zone = curr_req_ebus
                    added_card = auto_ebus_card
                    ebus_balance = FULL_BECKHOFF_DB[added_card]["ebus"]
                    power_balance = FULL_BECKHOFF_DB[added_card]["power"] 
                    if curr_req_power: power_zone = curr_req_power
                    
                    final_rack.append({
                        "Pozíció": real_pos, "Terminál": added_card.split(" (")[0], "Típus": "💡 Auto Leválasztás (Név)",
                        "E-Bus Zóna ⚡": ebus_zone, "E-Bus (mA)": int(ebus_balance),
                        "Terepi Zóna 🔌": power_zone, "Terepi (A)": round(power_balance, 2)
                    })
                    real_pos += 1; auto_added_count += 1; curr_req_power = ""

                if curr_req_power and curr_req_power != power_zone and not is_power_refresher:
                    power_zone = curr_req_power
                    added_card = auto_power_card
                    if FULL_BECKHOFF_DB[added_card]["ebus"] > 0: ebus_balance = FULL_BECKHOFF_DB[added_card]["ebus"]
                    else: ebus_balance += FULL_BECKHOFF_DB[added_card]["ebus"]
                    if FULL_BECKHOFF_DB[added_card]["power"] > 0: power_balance = FULL_BECKHOFF_DB[added_card]["power"]
                    else: power_balance += FULL_BECKHOFF_DB[added_card]["power"]
                    
                    final_rack.append({
                        "Pozíció": real_pos, "Terminál": added_card.split(" (")[0], "Típus": "💡 Auto Leválasztás (Név)",
                        "E-Bus Zóna ⚡": ebus_zone, "E-Bus (mA)": int(ebus_balance),
                        "Terepi Zóna 🔌": power_zone, "Terepi (A)": round(power_balance, 2)
                    })
                    real_pos += 1; auto_added_count += 1

                if not is_ebus_refresher and ebus_balance + card_ebus < 0:
                    added_card = auto_ebus_card
                    ebus_balance = FULL_BECKHOFF_DB[added_card]["ebus"]
                    power_balance = FULL_BECKHOFF_DB[added_card]["power"]
                    ebus_zone += " (Auto)"
                    final_rack.append({
                        "Pozíció": real_pos, "Terminál": added_card.split(" (")[0], "Típus": "⚠️ AUTO ADDED (E-Bus Hiány)",
                        "E-Bus Zóna ⚡": ebus_zone, "E-Bus (mA)": int(ebus_balance),
                        "Terepi Zóna 🔌": power_zone, "Terepi (A)": round(power_balance, 2)
                    })
                    real_pos += 1; auto_added_count += 1

                if not is_power_refresher and power_balance + card_power < 0:
                    added_card = auto_power_card
                    if FULL_BECKHOFF_DB[added_card]["ebus"] > 0: ebus_balance = FULL_BECKHOFF_DB[added_card]["ebus"]
                    else: ebus_balance += FULL_BECKHOFF_DB[added_card]["ebus"]
                    if FULL_BECKHOFF_DB[added_card]["power"] > 0: power_balance = FULL_BECKHOFF_DB[added_card]["power"]
                    else: power_balance += FULL_BECKHOFF_DB[added_card]["power"]
                    power_zone += " (Auto)"
                    final_rack.append({
                        "Pozíció": real_pos, "Terminál": added_card.split(" (")[0], "Típus": "⚠️ AUTO ADDED (Terepi Hiány)",
                        "E-Bus Zóna ⚡": ebus_zone, "E-Bus (mA)": int(ebus_balance),
                        "Terepi Zóna 🔌": power_zone, "Terepi (A)": round(power_balance, 2)
                    })
                    real_pos += 1; auto_added_count += 1

                if not is_ebus_refresher: ebus_balance += card_ebus
                if not is_power_refresher: power_balance += card_power

                final_rack.append({
                    "Pozíció": real_pos, 
                    "Terminál": t_type.split(" (")[0],
                    "Típus": "✅ Tápkártya" if is_ebus_refresher or is_power_refresher else "IO Modul",
                    "E-Bus Zóna ⚡": ebus_zone, 
                    "E-Bus (mA)": int(ebus_balance),
                    "Terepi Zóna 🔌": power_zone, 
                    "Terepi (A)": round(power_balance, 2)
                })
                real_pos += 1

        if final_rack:
            res_df = pd.DataFrame(final_rack)
            
            def style_auto_rows(row):
                if "⚠️ AUTO" in str(row["Típus"]): return ['background-color: #FFF3CD; color: #856404; font-weight: bold;'] * len(row)
                elif "💡 Auto Leválasztás" in str(row["Típus"]): return ['background-color: #CCE5FF; color: #004085; font-weight: bold;'] * len(row)
                elif "✅ Tápkártya" in str(row["Típus"]): return ['background-color: #D4EDDA; color: #155724; font-weight: bold;'] * len(row)
                return [''] * len(row)
                
            st.dataframe(res_df.style.apply(style_auto_rows, axis=1), use_container_width=True, hide_index=True)
            
            # ⚡ EL6070 LICENCKULCS FIGYELMEZTETÉS ÉS OKOS HOZZÁADÁS GOMB
            has_ek = False
            has_el6070 = False
            insert_sorszam = 11
            
            for idx, row in processed_rack.iterrows():
                t_type_str = str(row.get("Terminál típus", ""))
                if "EL6070" in t_type_str:
                    has_el6070 = True
                    break
                if t_type_str.startswith("EK") and not has_ek:
                    has_ek = True
                    insert_sorszam = int(row.get("Sorszám", 10)) + 1

            if has_ek and not has_el6070:
                st.write("")
                c_w1, c_w2 = st.columns([3, 1])
                with c_w1:
                    st.warning("⚠️ **Nem található EL6070 a konfigurációban. Szeretne hozzáadni?**")
                with c_w2:
                    if st.button("➕ EL6070 hozzáadása", use_container_width=True):
                        new_row = pd.DataFrame([{
                            "Sorszám": insert_sorszam,
                            "Terminál típus": "EL6070 (EtherCAT licenckulcs terminál)",
                            "Darabszám": 1,
                            "E-Bus Zóna": "",
                            "Terepi Zóna": ""
                        }])
                        st.session_state.beckhoff_rack_simple = pd.concat([edited_rack, new_row], ignore_index=True)
                        st.rerun()
            
            st.write("")
            c_m1, c_m2, c_m3 = st.columns(3)
            with c_m1: st.metric("Összes beépített kártya", f"{real_pos - 1} db")
            with c_m2: st.metric("Automatikusan befűzött táp", f"{auto_added_count} db", delta="Módosítva" if auto_added_count > 0 else "Tökéletes", delta_color="inverse" if auto_added_count > 0 else "normal")
            with c_m3:
                if auto_added_count == 0: st.success("✅ A manuális terved tökéletes!")
                else: st.warning(f"⚠️ A rendszer {auto_added_count} ponton rakott be okos leválasztót.")